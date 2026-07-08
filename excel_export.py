import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from database import get_all_professors


def build_excel() -> bytes:
    rows = get_all_professors()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "اساتید"
    ws.sheet_view.rightToLeft = True   

    headers = ["ردیف", "نام و نام‌خانوادگی", "کد ملی", "تلفن", "مدرک/رشته تحصیلی",
               "بازنشسته فنی‌حرفه‌ای", "گروه آموزشی", "روزها", "ساعت‌ها", "دروس", "تاریخ ثبت"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="B Nazanin", size=12)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, readingOrder=2)
        cell.border = border
    ws.row_dimensions[1].height = 30

    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    data_font = Font(name="B Nazanin", size=11)

    for i, row in enumerate(rows, 1):
        r = i + 1
        values = [i, row["full_name"], row["national_id"], row["phone"],
                  row["degree"], row["retired"],
                  row["department"], row["days"], row["hours"],
                  row["courses"], row["created_at"]]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True, readingOrder=2)
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[r].height = 25

    col_widths = [8, 22, 14, 14, 20, 16, 16, 28, 28, 38, 18]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
